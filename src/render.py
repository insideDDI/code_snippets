"""Render markdown to HTML using Jinja2 templates."""

import argparse
import json
from pathlib import Path
from typing import Any

import openpyxl
import yaml
from jinja2 import (
    Environment,
    FileSystemLoader
)


class RenderData:
    """Base class for data to be rendered."""

    def __init__(self, config_file: str = 'config.yaml'):
        """Initialize."""

        with open(Path('documents', 'source', config_file)) as infile:
            data = yaml.safe_load(infile.read())
        for k, v in data.items():
            setattr(self, k, v)


class Renderer:

    def __init__(
            self, extattrs: dict[str, Any] | None = None, config_file:
            str = 'config.yaml', ea_file: str | None = None
    ):
        """Initialize."""
        self.data = RenderData(config_file)

        if ea_file:
            self._ea_file = ea_file

            with open(Path('documents', 'source', self._ea_file)) as infile:
                self._extensible_attributes = yaml.safe_load(
                    infile.read()
                )

        else:
            self._extensible_attributes = extattrs

    def render_master(self):
        """Render markdown to HTML."""

        directory = 'extensible_attributes'

        _j_env = Environment(
            loader=FileSystemLoader(
                Path('jinja_templates', directory)
            )
        )
        template = Renderer._template = _j_env.get_template(
            'master.md'
        )

        results = template.render(data=self.data)
        return results

    def render_ea(self):
        """Render markdown to HTML."""

        directory = 'extensible_attributes'

        _j_env = Environment(
            loader=FileSystemLoader(
                Path('jinja_templates', directory)
            )
        )
        results: dict = {}

        for attr_name, attr_data in self._extensible_attributes.items():
            template = Renderer._template = _j_env.get_template(
                'extensible_attribute_description.md'
            )
            print(attr_name)
            results[attr_name] = template.render(
                attribute_name=attr_name,
                attribute_description=attr_data['comment'],
                attribute_data_type=attr_data['type'],
                attribute_flags=attr_data['flags'],
                attribute_ddi_objects=attr_data.get('allowed_objects'),
                attribute_enum_values=attr_data.get('list_values')

            )

        self.results = results

    def write_results(self):
        for attr_name, attr_data in self._extensible_attributes.items():
            path = Path(
                'documents',
                'output',
                f'{Renderer.snake_case_fold(attr_name)}.md'
            )
            print(path.absolute())
            with open(
                    path,
                    'w'
            ) as outfile:
                outfile.write(self.results[attr_name])

    @staticmethod
    def snake_case_fold(s: str) -> str:
        """Convert string to snake case."""
        return (str(s)
                .replace(' ', '_').casefold())


class ExtensibleAttributesExtract:

    def converter(excel_filename, outfile='output'):
        wb = openpyxl.load_workbook(excel_filename)
        # ws = wb['NetworkHierarchy']
        # inrow = 3
        output = dict(
            columns=dict(),
            address_type=dict(p='Network purpose', s='Site bound network'),
            choices=dict(),
            regions=dict(extattr='Region', choices=list())
        )

        # parse lines 3 through end
        # for row in range(inrow, ws.max_row + 1):
        #     incol = 1
        #     # skip empty lines
        #     if ws.cell(row=row, column=1).value:
        #         print(row)
        #         level0 = output['choices']
        #         k1 = ws.cell(row=row, column=incol).value
        #         if str(k1) not in level0.keys():
        #             level1 = dict()
        #             level0[str(k1)] = level1
        #         else:
        #             level1 = level0[str(k1)]
        #
        #         k2 = ws.cell(row=row, column=incol + 1).value
        #         if str(k2) not in level1.keys():
        #             level2 = dict()
        #             level1[str(k2)] = level2
        #         else:
        #             level2 = level1[str(k2)]
        #
        #         k3 = ws.cell(row=row, column=incol + 2).value
        #         if str(k3) not in level2.keys():
        #             level3 = dict()
        #             level2[str(k3)] = level3
        #         else:
        #             level3 = level2[str(k3)]
        #         address_type = ws.cell(row=row, column=incol + 3).value
        #
        #         level3[address_type] = dict(
        #             p=ws.cell(row=row, column=incol + 4).value.split(','),
        #             s=ws.cell(row=row, column=incol + 5).value
        #             )
        ws = wb['Extensible Attributes']
        column_headers = list()
        for incol in range(1, ws.max_column + 1):
            colname = ws.cell(row=1, column=incol).value
            if colname and colname not in column_headers:
                column_headers.append(colname)
        extattrs = dict()
        for inrow in range(2, ws.max_row + 1):
            fieldname = ws.cell(row=inrow, column=2).value
            if fieldname:

                extattr = ws.cell(row=inrow, column=3).value
                extattrs[extattr] = dict()
                output['columns'][fieldname] = extattr
                # process comment
                extattrs[extattr]['comment'] = ws.cell(
                    row=inrow, column=4
                ).value
                extattr_type = ws.cell(row=inrow, column=5).value.upper()
                # parse the worksheet with list values
                extattrs[extattr]['type'] = extattr_type
                # If the extattr is a list, then read excel file tables and
                # return list values into 'list_values' list
                if extattr_type == 'ENUM':
                    if fieldname not in wb.sheetnames:
                        raise ValueError(f'{fieldname} not in wb.sheetnames')
                    extattrs[extattr]['list_values'] = list()
                    wsea = wb[fieldname]
                    for row in range(2, wsea.max_row + 1):
                        value = wsea.cell(row=row, column=1).value
                        if value:
                            extattrs[extattr]['list_values'].append(dict(
                                value=value, comment=wsea.cell(
                                    row=row, column=2
                                ).value

                # process flags
                # format: CGILMPRSV
                # V = multiple
                # M - mandatory
                # L - listed (recommended for object)
                flags = dict()
                # required flag
                f = ws.cell(row=inrow, column=9).value
                if f == 'yes':
                    flags['required'] = "M"
                elif f == 'recommended':
                    flags['required'] = "L"
                else:
                    flags['required'] = ''
                # Inherit flag
                f = ws.cell(row=inrow, column=6).value
                if f == 'yes':
                    flags['inherit'] = 'I'
                else:
                    flags['inherit'] = ''
                # Multiple values flag
                f = ws.cell(row=inrow, column=7).value
                if f == 'yes':
                    flags['multiple'] = 'V'
                else:
                    flags['multiple'] = ''

                extattrs[extattr]['flags'] = '{}{}{}'.format(
                    flags['inherit'],
                    flags['required'],
                    flags['multiple']
                )

                # process allowed_objects
                startcol = 15
                allowed_objects = list()
                for incol in range(startcol, ws.max_column + 1):
                    value = ws.cell(row=inrow, column=incol).value
                    if value:
                        allowed_objects.append(value)
                if len(allowed_objects) > 0:
                    extattrs[extattr]['allowed_objects'] = allowed_objects

        with open(f'{outfile}.json', 'w') as outf:
            outf.write(json.dumps(output, indent=4))
        with open(f'{outfile}-no-indent.json', 'w') as outf:
            outf.write(json.dumps(output))
        with open(f'{outfile}-eas.json', 'w') as outf:
            outf.write(json.dumps(extattrs, indent=4))
        with open(f'{outfile}-eas.yaml', 'w') as outf:
            outf.write(yaml.dump(extattrs, indent=4, width=100))

        return extattrs


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument(
        '--file',
        help='Excel file to process',
        required=True
    )
    parser.add_argument('--config', help='Config file', default='config.yaml')
    args = parser.parse_args()

    ea = ExtensibleAttributesExtract.converter(args.file)
    r = Renderer(ea, args.config)
    r.render_ea()
    r.write_results()
